from datetime import datetime, timedelta
import os
import zipfile

from openpyxl import Workbook, load_workbook
import pytest

from flows.financeiro_hoje.workbook import WorkbookInvalid, consolidate, validate


HEADERS = ["Numero", "SituacaoOS", "Fluxo", "Cliente"]


def make_book(path, rows, headers=HEADERS, sheet_name="Consulta"):
    book = Workbook()
    sheet = book.active
    sheet.title = sheet_name
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    book.save(path)


def test_consolidacao_apenas_concatena_e_preserva_ordem(tmp_path):
    loga, acerta, out = tmp_path / "loga.xlsx", tmp_path / "acerta.xlsx", tmp_path / "out.xlsx"
    make_book(loga, [[1, "Na Fila", "#1.25", "A"]])
    make_book(acerta, [[1, "Concluido", "#1.25", "A"], [2, "Na Fila", "#3.10", "B"]])

    result = consolidate(loga, acerta, out)

    sheet = load_workbook(out, read_only=True)["Consulta"]
    assert list(sheet.values) == [
        tuple(HEADERS),
        (1, "Na Fila", "#1.25", "A"),
        (1, "Concluido", "#1.25", "A"),
        (2, "Na Fila", "#3.10", "B"),
    ]
    assert result.rows == 3


def test_validacao_rejeita_fonte_sem_linhas(tmp_path):
    path = tmp_path / "empty.xlsx"
    make_book(path, [])

    with pytest.raises(WorkbookInvalid, match="sem linhas"):
        validate(path)


def test_validacao_le_todo_workbook_e_retorna_metadados(tmp_path):
    path = tmp_path / "report.xlsx"
    make_book(path, [[1, "Na Fila", "#1.25", "A"], [2, "Concluido", "#3.10", "B"]])

    artifact = validate(path)

    assert artifact.path == path
    assert artifact.headers == tuple(HEADERS)
    assert artifact.rows == 2
    assert artifact.size == path.stat().st_size
    assert len(artifact.sha256) == 64


def test_validacao_rejeita_cabecalho_obrigatorio_ausente(tmp_path):
    path = tmp_path / "missing-header.xlsx"
    make_book(path, [[1, "Na Fila", "A"]], headers=["Numero", "SituacaoOS", "Cliente"])

    with pytest.raises(WorkbookInvalid, match="Fluxo"):
        validate(path)


def test_validacao_rejeita_arquivo_que_nao_e_zip_openxml(tmp_path):
    path = tmp_path / "invalid.xlsx"
    path.write_text("nao e um arquivo xlsx", encoding="utf-8")

    with pytest.raises(WorkbookInvalid, match="ZIP"):
        validate(path)


def test_validacao_rejeita_zip_que_nao_e_pacote_openxml(tmp_path):
    path = tmp_path / "not-openxml.xlsx"
    with zipfile.ZipFile(path, "w") as archive:
        archive.writestr("plain.txt", "not an XLSX package")

    with pytest.raises(WorkbookInvalid, match="ZIP"):
        validate(path)


def test_validacao_rejeita_linha_totalmente_vazia(tmp_path):
    path = tmp_path / "blank-row.xlsx"
    make_book(path, [[None, None, None, None]])

    with pytest.raises(WorkbookInvalid, match="sem linhas"):
        validate(path)


def test_validacao_traduz_xml_malformado_e_libera_arquivo(tmp_path):
    path = tmp_path / "malformed-sheet.xlsx"
    replacement = tmp_path / "replacement.xlsx"
    make_book(path, [[1, "Na Fila", "#1.25", "A"]])

    with zipfile.ZipFile(path) as source, zipfile.ZipFile(replacement, "w") as target:
        for item in source.infolist():
            content = source.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                content = b"<worksheet><sheetData>"
            target.writestr(item, content)
    replacement.replace(path)

    with pytest.raises(WorkbookInvalid, match="OpenXML"):
        validate(path)

    with path.open("rb") as reopened:
        assert reopened.read(2) == b"PK"
    path.unlink()
    assert not path.exists()

def test_consolidacao_normaliza_cabecalhos_equivalentes_routerbox(tmp_path):
    loga = tmp_path / "loga.xlsx"
    acerta = tmp_path / "acerta.xlsx"
    out = tmp_path / "out.xlsx"
    loga_headers = [
        "Numero", "Data_Abertura", "Data_Hora", "Designado_Para",
        "SituacaoOS", "Causa OS", "Fluxo", "Grupo Cobran\u00e7a", "Causa Atend",
    ]
    acerta_headers = [
        "Numero", "DataAbertura", "Agendamento", "DesignadoPara",
        "SituacaoOS", "Causa", "Fluxo", "GrupoCobranca", "CausaAtend",
    ]
    make_book(loga, [[1, "28/07/2026", "08:00", "Equipe", "Concluido", "Ativacao", "#1", "A", "Ok"]], headers=loga_headers)
    make_book(acerta, [[2, "28/07/2026", "09:00", "Equipe", "Concluido", "Reparo", "#2", "B", "Ok"]], headers=acerta_headers)

    result = consolidate(loga, acerta, out)

    sheet = load_workbook(out, read_only=True)["Consulta"]
    assert tuple(next(sheet.values)) == tuple(loga_headers)
    assert result.rows == 2


def test_validacao_rejeita_arquivo_mais_antigo_que_a_execucao(tmp_path):
    path = tmp_path / "old.xlsx"
    make_book(path, [[1, "Na Fila", "#1.25", "A"]])
    old = (datetime.now() - timedelta(minutes=5)).timestamp()
    os.utime(path, (old, old))

    with pytest.raises(WorkbookInvalid, match="execucao"):
        validate(path, created_after=datetime.now() - timedelta(minutes=1))


def test_consolidacao_rejeita_fontes_com_cabecalhos_diferentes(tmp_path):
    loga, acerta, out = tmp_path / "loga.xlsx", tmp_path / "acerta.xlsx", tmp_path / "out.xlsx"
    make_book(loga, [[1, "Na Fila", "#1.25", "A"]])
    make_book(acerta, [[1, "Na Fila", "#1.25", "A"]], headers=HEADERS + ["Extra"])

    with pytest.raises(WorkbookInvalid, match="cabecalhos"):
        consolidate(loga, acerta, out)
