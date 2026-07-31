from datetime import datetime
import json
import os
from pathlib import Path
from shutil import rmtree
import subprocess
import sys
import tempfile
from types import SimpleNamespace
import unittest
from unittest.mock import patch
from zoneinfo import ZoneInfo

from openpyxl import Workbook

from flows.common.locks import file_lock
from flows.financeiro_medicao import bundle, runner, schedule
from flows.financeiro_medicao.cycles import window_for
from flows.financeiro_medicao.workbook import (
    REQUIRED_HEADERS,
    SHEET_NAME,
)
import worker


CANONICAL_ROW = (
    "123456",
    "11/07/2026 08:00:00",
    "PROTO-123",
    "Cliente Exemplo",
    "Maria da Silva",
    "Sao Paulo",
    "Centro",
    "Usuario Final",
    "Executor Exemplo",
    "15/07/2026 12:30:00",
    "Empresa Exemplo LTDA",
    "Empresa Exemplo",
    "Financeiro",
    "Medicao",
    "Pagamento",
    "Sem causa",
    "Encerrado",
    "Nao",
    "",
    125.50,
)


class FinanceiroMedicaoScheduleTests(unittest.TestCase):
    def setUp(self):
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name) / "financeiro_medicao"
        self.runtime = self.root / "runtime"
        self.runtime.mkdir(parents=True)
        self.timezone = ZoneInfo("America/Sao_Paulo")
        self.scheduled_for = datetime(
            2026,
            7,
            30,
            0,
            1,
            tzinfo=self.timezone,
        )
        self.settings = SimpleNamespace(
            runtime_root=self.root,
            schedule_enabled=True,
            schedule_hour=0,
            schedule_minute=1,
            timezone="America/Sao_Paulo",
        )

    def tearDown(self):
        self.temporary.cleanup()

    @property
    def signal(self):
        return self.runtime / "schedule.signal.json"

    @property
    def watermark(self):
        return self.runtime / "schedule-watermark.json"

    def _request(self):
        schedule.request_run(
            self.settings,
            self.scheduled_for,
            now=self.scheduled_for,
        )

    def _claim(self, suffix, claimed_at="2026-07-30T00:01:00-03:00"):
        event_id, _ = runner.scheduled_event_identity(
            self.scheduled_for
        )
        claim = self.runtime / f".schedule.signal.json.claimed.{suffix}"
        claim.write_text(
            json.dumps(
                {
                    "schema_version": 1,
                    "event_id": event_id,
                    "scheduled_for": self.scheduled_for.isoformat(),
                    "attempt": 0,
                    "next_attempt_at": self.scheduled_for.isoformat(),
                    "claimed_at": claimed_at,
                }
            ),
            encoding="utf-8",
        )
        return claim

    def test_retry_crossing_midnight_completes_original_day_only(self):
        july_10 = self.scheduled_for.replace(day=10)
        july_11_now = july_10.replace(day=11, minute=3)

        attempts = 0

        def run_once(**_arguments):
            nonlocal attempts
            attempts += 1
            if attempts == 1:
                return {"success": False, "error_code": "LOCKED"}
            (self.root / "done.json").write_text(
                json.dumps(
                    {
                        "success": True,
                        "finished_at": july_11_now.isoformat(),
                    }
                ),
                encoding="utf-8",
            )
            return {"success": True, "error_code": ""}

        with patch.object(
            schedule.runner,
            "run_once",
            side_effect=run_once,
        ) as scheduled_runner:
            schedule.request_run(
                self.settings,
                july_10,
                now=july_10,
            )
            self.assertFalse(
                schedule.run_signal_if_due(
                    self.settings,
                    now=july_10,
                )
            )
            self.assertTrue(
                schedule.run_signal_if_due(
                    self.settings,
                    now=july_11_now,
                )
            )

        watermark = json.loads(
            self.watermark.read_text(encoding="utf-8")
        )
        expected_event_id, _ = runner.scheduled_event_identity(july_10)
        self.assertEqual(
            [
                call.kwargs["event_id"]
                for call in scheduled_runner.call_args_list
            ],
            [expected_event_id, expected_event_id],
        )
        self.assertEqual(watermark["local_date"], "2026-07-10")
        self.assertEqual(
            schedule.next_event_at(july_11_now, self.settings),
            july_10.replace(day=11),
        )

    def test_live_event_owner_defers_claim_without_hot_loop(self):
        claim = self._claim("c" * 32)
        event_id, _ = runner.scheduled_event_identity(
            self.scheduled_for
        )
        owner_lock = (
            self.runtime / "events" / f"{event_id}.lock"
        )
        now = self.scheduled_for.replace(minute=10)

        with file_lock(owner_lock, wait_seconds=0), patch.object(
            schedule.runner,
            "run_once",
        ) as run_once:
            next_event = schedule.next_event_at(now, self.settings)
            result = schedule.run_signal_if_due(
                self.settings,
                now=now,
            )

        self.assertGreater(next_event, now)
        self.assertIsNone(result)
        self.assertTrue(claim.exists())
        run_once.assert_not_called()

    def test_event_owner_reparse_directory_is_rejected(self):
        claim = self._claim("e" * 32)
        events_dir = self.runtime / "events"
        events_dir.mkdir()
        original_lstat = Path.lstat

        def report_reparse(path):
            metadata = original_lstat(path)
            if Path(path) == events_dir:
                return SimpleNamespace(
                    st_mode=metadata.st_mode,
                    st_dev=metadata.st_dev,
                    st_ino=metadata.st_ino,
                    st_size=metadata.st_size,
                    st_mtime_ns=metadata.st_mtime_ns,
                    st_ctime_ns=metadata.st_ctime_ns,
                    st_file_attributes=0x400,
                )
            return metadata

        with patch.object(
            Path,
            "lstat",
            new=report_reparse,
        ), patch.object(
            schedule.runner,
            "run_once",
        ) as run_once:
            result = schedule.run_signal_if_due(
                self.settings,
                now=self.scheduled_for,
            )

        self.assertIsNone(result)
        self.assertTrue(claim.exists())
        run_once.assert_not_called()

    def test_event_owner_directory_swap_is_rejected(self):
        claim = self._claim("f" * 32)
        events_dir = self.runtime / "events"
        events_dir.mkdir()
        original_lstat = Path.lstat
        observations = 0

        def report_swap(path):
            nonlocal observations
            metadata = original_lstat(path)
            if Path(path) == events_dir:
                observations += 1
                if observations > 1:
                    return SimpleNamespace(
                        st_mode=metadata.st_mode,
                        st_dev=metadata.st_dev,
                        st_ino=metadata.st_ino + observations,
                        st_nlink=metadata.st_nlink,
                        st_size=metadata.st_size,
                        st_mtime_ns=metadata.st_mtime_ns,
                        st_ctime_ns=metadata.st_ctime_ns,
                        st_file_attributes=0,
                    )
            return metadata

        with patch.object(
            Path,
            "lstat",
            new=report_swap,
        ), patch.object(
            schedule.runner,
            "run_once",
        ) as run_once:
            result = schedule.run_signal_if_due(
                self.settings,
                now=self.scheduled_for,
            )

        self.assertIsNone(result)
        self.assertTrue(claim.exists())
        run_once.assert_not_called()

    def test_terminal_journal_is_reconciled_before_rerun(self):
        claim = self._claim("d" * 32)
        event_id, _ = runner.scheduled_event_identity(
            self.scheduled_for
        )
        payload = {
            "success": False,
            "error_code": "AUTH_EXPIRED",
            "run_id": None,
            "cycle_id": "2026-07-11--2026-08-10",
            "mode": "current",
            "started_at": self.scheduled_for.isoformat(),
            "finished_at": self.scheduled_for.isoformat(),
            "next_scheduled_for": None,
        }
        journal = runner.event_result_path(self.root, event_id)
        journal.parent.mkdir(parents=True)
        journal.write_text(
            json.dumps(
                {
                    "schema_version": 1,
                    "event_id": event_id,
                    "scheduled_for": self.scheduled_for.isoformat(),
                    "payload": payload,
                }
            ),
            encoding="utf-8",
        )

        with patch.object(
            schedule.runner,
            "run_once",
            side_effect=AssertionError("must not rerun"),
        ) as run_once:
            result = schedule.run_signal_if_due(
                self.settings,
                now=self.scheduled_for.replace(minute=10),
            )

        self.assertFalse(result)
        self.assertFalse(claim.exists())
        self.assertTrue(
            self.watermark.exists(),
            "terminal journal must be finalized without rerunning",
        )
        self.assertEqual(
            json.loads(self.watermark.read_text(encoding="utf-8"))[
                "outcome"
            ],
            "terminal",
        )
        run_once.assert_not_called()

    def test_published_bundle_is_reconciled_before_rerun(self):
        event_id, run_id = runner.scheduled_event_identity(
            self.scheduled_for
        )
        source = Path(self.temporary.name) / "source.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = SHEET_NAME
        worksheet.append(list(REQUIRED_HEADERS))
        worksheet.append(list(CANONICAL_ROW))
        workbook.save(source)
        workbook.close()
        published = bundle.build_bundle(
            runtime_root=self.root,
            source=source,
            window=window_for(self.scheduled_for.date()),
            scheduled_for=self.scheduled_for,
            started_at=self.scheduled_for,
            finished_at=self.scheduled_for,
            image_revision="test-revision",
            run_id=run_id,
        )
        claim = self._claim("e" * 32)

        with patch.object(
            schedule.runner,
            "run_once",
            side_effect=AssertionError("must not capture again"),
        ) as run_once:
            result = schedule.run_signal_if_due(
                self.settings,
                now=self.scheduled_for.replace(minute=10),
            )

        self.assertTrue(result)
        self.assertEqual(published.name, run_id)
        self.assertFalse(claim.exists())
        self.assertTrue(
            runner.event_result_path(self.root, event_id).is_file()
        )
        self.assertEqual(
            json.loads(self.watermark.read_text(encoding="utf-8"))[
                "outcome"
            ],
            "success",
        )
        run_once.assert_not_called()

    def test_success_journal_survives_bundle_consumption_via_receipt(self):
        event_id, run_id = runner.scheduled_event_identity(
            self.scheduled_for
        )
        source = Path(self.temporary.name) / "source.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = SHEET_NAME
        worksheet.append(list(REQUIRED_HEADERS))
        worksheet.append(list(CANONICAL_ROW))
        workbook.save(source)
        workbook.close()
        published = bundle.build_bundle(
            runtime_root=self.root,
            source=source,
            window=window_for(self.scheduled_for.date()),
            scheduled_for=self.scheduled_for,
            started_at=self.scheduled_for,
            finished_at=self.scheduled_for,
            image_revision="test-revision",
            run_id=run_id,
        )
        payload = runner.reconcile_published_event(
            self.settings,
            event_id,
            self.scheduled_for,
        )
        consumed = self.root / "published" / run_id
        consumed.parent.mkdir()
        os.replace(published, consumed)

        receipt = runner.event_receipt_path(
            self.root,
            event_id,
        )
        recovered = runner.read_event_result(
            self.root,
            event_id,
            self.scheduled_for,
        )

        self.assertTrue(payload["success"])
        self.assertEqual(recovered, payload)
        self.assertTrue(receipt.is_file())
        if os.name == "posix":
            self.assertEqual(
                os.stat(receipt).st_mode & 0o777,
                0o600,
            )
            self.assertEqual(
                os.stat(receipt.parent).st_mode & 0o777,
                0o700,
            )

    def test_receipt_and_journal_without_independent_proof_are_rejected(self):
        event_id, run_id = runner.scheduled_event_identity(
            self.scheduled_for
        )
        source = Path(self.temporary.name) / "source.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = SHEET_NAME
        worksheet.append(list(REQUIRED_HEADERS))
        worksheet.append(list(CANONICAL_ROW))
        workbook.save(source)
        workbook.close()
        published = bundle.build_bundle(
            runtime_root=self.root,
            source=source,
            window=window_for(self.scheduled_for.date()),
            scheduled_for=self.scheduled_for,
            started_at=self.scheduled_for,
            finished_at=self.scheduled_for,
            image_revision="test-revision",
            run_id=run_id,
        )
        runner.reconcile_published_event(
            self.settings,
            event_id,
            self.scheduled_for,
        )
        consumed = Path(self.temporary.name) / "consumed" / run_id
        consumed.parent.mkdir()
        os.replace(published, consumed)
        proof = self.root / "runtime" / "proofs" / run_id
        if proof.exists():
            rmtree(proof)

        with self.assertRaisesRegex(
            ValueError,
            "Invalid event result journal",
        ):
            runner.read_event_result(
                self.root,
                event_id,
                self.scheduled_for,
            )

    def test_prepared_proof_does_not_reconcile_and_retry_publishes(self):
        event_id, run_id = runner.scheduled_event_identity(
            self.scheduled_for
        )
        source = Path(self.temporary.name) / "source.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = SHEET_NAME
        worksheet.append(list(REQUIRED_HEADERS))
        worksheet.append(list(CANONICAL_ROW))
        workbook.save(source)
        workbook.close()
        original_replace = os.replace

        def crash_before_inbox(source_path, destination):
            if Path(destination).parent == self.root / "inbox":
                raise OSError("crash before inbox publication")
            return original_replace(source_path, destination)

        arguments = dict(
            runtime_root=self.root,
            source=source,
            window=window_for(self.scheduled_for.date()),
            scheduled_for=self.scheduled_for,
            started_at=self.scheduled_for,
            finished_at=self.scheduled_for,
            image_revision="test-revision",
            run_id=run_id,
        )
        with patch.object(
            bundle.os,
            "replace",
            side_effect=crash_before_inbox,
        ), self.assertRaisesRegex(OSError, "crash before inbox"):
            bundle.build_bundle(**arguments)

        reconciled = runner.reconcile_published_event(
            self.settings,
            event_id,
            self.scheduled_for,
        )
        published = bundle.build_bundle(**arguments)

        self.assertIsNone(reconciled)
        self.assertEqual(published, self.root / "inbox" / run_id)
        self.assertTrue(published.is_dir())
        self.assertIsNotNone(
            bundle.inspect_committed_publication(
                runtime_root=self.root,
                run_id=run_id,
                window=window_for(self.scheduled_for.date()),
                scheduled_for=self.scheduled_for,
            )
        )

    def test_committed_bundle_consumed_immediately_reconciles_once(self):
        event_id, run_id = runner.scheduled_event_identity(
            self.scheduled_for
        )
        source = Path(self.temporary.name) / "source.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = SHEET_NAME
        worksheet.append(list(REQUIRED_HEADERS))
        worksheet.append(list(CANONICAL_ROW))
        workbook.save(source)
        workbook.close()
        consumed = self.root / "published" / run_id
        consumed.parent.mkdir()
        original_replace = os.replace

        def consume_canonical(source_path, destination):
            result = original_replace(source_path, destination)
            if Path(destination) == self.root / "inbox" / run_id:
                original_replace(destination, consumed)
            return result

        arguments = dict(
            runtime_root=self.root,
            source=source,
            window=window_for(self.scheduled_for.date()),
            scheduled_for=self.scheduled_for,
            started_at=self.scheduled_for,
            finished_at=self.scheduled_for,
            image_revision="test-revision",
            run_id=run_id,
        )
        with patch.object(
            bundle.os,
            "replace",
            side_effect=consume_canonical,
        ):
            bundle.build_bundle(**arguments)

        committed = bundle.inspect_committed_publication(
            runtime_root=self.root,
            run_id=run_id,
            window=window_for(self.scheduled_for.date()),
            scheduled_for=self.scheduled_for,
        )
        payload = runner.reconcile_published_event(
            self.settings,
            event_id,
            self.scheduled_for,
        )
        repeated = bundle.build_bundle(**arguments)

        self.assertIsNotNone(committed)
        self.assertTrue(payload["success"])
        self.assertEqual(repeated, bundle.publication_proof_path(
            self.root,
            run_id,
        ))
        self.assertTrue(consumed.is_dir())
        self.assertFalse((self.root / "inbox" / run_id).exists())

    def test_committed_pending_bundle_is_finalized_before_reconcile(self):
        event_id, run_id = runner.scheduled_event_identity(
            self.scheduled_for
        )
        source = Path(self.temporary.name) / "source.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = SHEET_NAME
        worksheet.append(list(REQUIRED_HEADERS))
        worksheet.append(list(CANONICAL_ROW))
        workbook.save(source)
        workbook.close()
        canonical = self.root / "inbox" / run_id
        pending = bundle.publication_pending_path(
            self.root,
            run_id,
        )
        original_replace = os.replace

        def crash_before_canonical(source_path, destination):
            if Path(destination) == canonical:
                raise OSError("crash before canonical rename")
            return original_replace(source_path, destination)

        with patch.object(
            bundle.os,
            "replace",
            side_effect=crash_before_canonical,
        ), self.assertRaisesRegex(OSError, "canonical rename"):
            bundle.build_bundle(
                runtime_root=self.root,
                source=source,
                window=window_for(self.scheduled_for.date()),
                scheduled_for=self.scheduled_for,
                started_at=self.scheduled_for,
                finished_at=self.scheduled_for,
                image_revision="test-revision",
                run_id=run_id,
            )

        self.assertTrue(pending.is_dir())
        self.assertFalse(canonical.exists())
        payload = runner.reconcile_published_event(
            self.settings,
            event_id,
            self.scheduled_for,
        )

        self.assertTrue(payload["success"])
        self.assertTrue(canonical.is_dir())
        self.assertFalse(pending.exists())

    def test_inconsistent_receipt_cannot_authorize_success_journal(self):
        event_id, run_id = runner.scheduled_event_identity(
            self.scheduled_for
        )
        source = Path(self.temporary.name) / "source.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = SHEET_NAME
        worksheet.append(list(REQUIRED_HEADERS))
        worksheet.append(list(CANONICAL_ROW))
        workbook.save(source)
        workbook.close()
        published = bundle.build_bundle(
            runtime_root=self.root,
            source=source,
            window=window_for(self.scheduled_for.date()),
            scheduled_for=self.scheduled_for,
            started_at=self.scheduled_for,
            finished_at=self.scheduled_for,
            image_revision="test-revision",
            run_id=run_id,
        )
        runner.reconcile_published_event(
            self.settings,
            event_id,
            self.scheduled_for,
        )
        receipt = runner.event_receipt_path(
            self.root,
            event_id,
        )
        receipt_payload = json.loads(
            receipt.read_text(encoding="utf-8")
        )
        receipt_payload["workbook_sha256"] = "0" * 64
        receipt.write_text(
            json.dumps(receipt_payload),
            encoding="utf-8",
        )
        consumed = self.root / "published" / run_id
        consumed.parent.mkdir()
        os.replace(published, consumed)

        with self.assertRaisesRegex(
            ValueError,
            "Invalid event result journal",
        ):
            runner.read_event_result(
                self.root,
                event_id,
                self.scheduled_for,
            )

    def test_receipt_alone_reconstructs_journal_after_bundle_moves(self):
        event_id, run_id = runner.scheduled_event_identity(
            self.scheduled_for
        )
        source = Path(self.temporary.name) / "source.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = SHEET_NAME
        worksheet.append(list(REQUIRED_HEADERS))
        worksheet.append(list(CANONICAL_ROW))
        workbook.save(source)
        workbook.close()
        published = bundle.build_bundle(
            runtime_root=self.root,
            source=source,
            window=window_for(self.scheduled_for.date()),
            scheduled_for=self.scheduled_for,
            started_at=self.scheduled_for,
            finished_at=self.scheduled_for,
            image_revision="test-revision",
            run_id=run_id,
        )
        details = bundle.inspect_published_bundle(
            runtime_root=self.root,
            run_id=run_id,
            window=window_for(self.scheduled_for.date()),
            scheduled_for=self.scheduled_for,
        )
        runner._write_success_receipt(
            self.root,
            event_id,
            self.scheduled_for,
            details,
        )
        consumed = self.root / "published" / run_id
        consumed.parent.mkdir()
        os.replace(published, consumed)

        payload = runner.reconcile_published_event(
            self.settings,
            event_id,
            self.scheduled_for,
        )

        self.assertTrue(payload["success"])
        self.assertTrue(
            runner.event_result_path(
                self.root,
                event_id,
            ).is_file()
        )

    def test_reconciliation_uses_captured_manifest_without_path_reread(self):
        event_id, run_id = runner.scheduled_event_identity(
            self.scheduled_for
        )
        source = Path(self.temporary.name) / "source.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = SHEET_NAME
        worksheet.append(list(REQUIRED_HEADERS))
        worksheet.append(list(CANONICAL_ROW))
        workbook.save(source)
        workbook.close()
        bundle.build_bundle(
            runtime_root=self.root,
            source=source,
            window=window_for(self.scheduled_for.date()),
            scheduled_for=self.scheduled_for,
            started_at=self.scheduled_for,
            finished_at=self.scheduled_for,
            image_revision="test-revision",
            run_id=run_id,
        )
        original_read_text = Path.read_text

        def reject_manifest_reread(path, *args, **kwargs):
            if Path(path).name == "manifest.json":
                raise AssertionError("manifest path reread")
            return original_read_text(path, *args, **kwargs)

        with patch.object(
            Path,
            "read_text",
            new=reject_manifest_reread,
        ):
            payload = runner.reconcile_published_event(
                self.settings,
                event_id,
                self.scheduled_for,
            )

        self.assertTrue(payload["success"])

    def test_corrupt_journal_without_bundle_is_quarantined_then_rerun(self):
        claim = self._claim("f" * 32)
        event_id, _ = runner.scheduled_event_identity(
            self.scheduled_for
        )
        journal = runner.event_result_path(self.root, event_id)
        journal.parent.mkdir(parents=True, exist_ok=True)
        journal.write_text("{not-json", encoding="utf-8")
        original_replace = os.replace
        renames = []

        def record_replace(source, destination):
            source = Path(source)
            destination = Path(destination)
            if source == journal:
                renames.append((source, destination))
            original_replace(source, destination)

        with patch.object(
            schedule.os,
            "replace",
            side_effect=record_replace,
        ), patch.object(
            schedule.runner,
            "run_once",
            return_value={"success": True, "error_code": ""},
        ) as run_once:
            result = schedule.run_signal_if_due(
                self.settings,
                now=self.scheduled_for.replace(minute=10),
            )

        self.assertTrue(result)
        self.assertFalse(claim.exists())
        self.assertFalse(journal.exists())
        self.assertEqual(len(renames), 1)
        self.assertEqual(renames[0][0], journal)
        self.assertEqual(renames[0][1].parent, journal.parent)
        self.assertIn(".corrupt.", renames[0][1].name)
        self.assertEqual(
            run_once.call_args.kwargs["event_id"],
            event_id,
        )
        self.assertEqual(
            run_once.call_args.kwargs["scheduled_for"],
            self.scheduled_for,
        )
        watermark = json.loads(
            self.watermark.read_text(encoding="utf-8")
        )
        self.assertNotEqual(
            watermark.get("error_code"),
            "CONFIG_INVALID",
        )

    def test_corrupt_journal_recovers_valid_bundle_before_rerun(self):
        event_id, run_id = runner.scheduled_event_identity(
            self.scheduled_for
        )
        source = Path(self.temporary.name) / "source.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = SHEET_NAME
        worksheet.append(list(REQUIRED_HEADERS))
        worksheet.append(list(CANONICAL_ROW))
        workbook.save(source)
        workbook.close()
        bundle.build_bundle(
            runtime_root=self.root,
            source=source,
            window=window_for(self.scheduled_for.date()),
            scheduled_for=self.scheduled_for,
            started_at=self.scheduled_for,
            finished_at=self.scheduled_for,
            image_revision="test-revision",
            run_id=run_id,
        )
        journal = runner.event_result_path(self.root, event_id)
        journal.parent.mkdir(parents=True, exist_ok=True)
        journal.write_text("{not-json", encoding="utf-8")
        claim = self._claim("1" * 32)

        with patch.object(
            schedule.runner,
            "run_once",
            side_effect=AssertionError("must not capture again"),
        ) as run_once:
            result = schedule.run_signal_if_due(
                self.settings,
                now=self.scheduled_for.replace(minute=10),
            )

        self.assertTrue(result)
        self.assertFalse(claim.exists())
        self.assertTrue(
            runner.read_event_result(
                self.root,
                event_id,
                self.scheduled_for,
            )["success"]
        )
        run_once.assert_not_called()

    def test_corrupt_journal_with_bundle_collision_is_terminal(self):
        event_id, run_id = runner.scheduled_event_identity(
            self.scheduled_for
        )
        published = self.root / "inbox" / run_id
        published.mkdir(parents=True)
        (published / "manifest.json").write_text(
            "{}",
            encoding="utf-8",
        )
        journal = runner.event_result_path(self.root, event_id)
        journal.parent.mkdir(parents=True, exist_ok=True)
        journal.write_text("{not-json", encoding="utf-8")
        claim = self._claim("2" * 32)

        with patch.object(
            schedule.runner,
            "run_once",
            side_effect=AssertionError("must not capture collision"),
        ) as run_once:
            result = schedule.run_signal_if_due(
                self.settings,
                now=self.scheduled_for.replace(minute=10),
            )

        self.assertFalse(result)
        self.assertFalse(claim.exists())
        watermark = json.loads(
            self.watermark.read_text(encoding="utf-8")
        )
        self.assertEqual(
            watermark["error_code"],
            "BUNDLE_COLLISION",
        )
        run_once.assert_not_called()

    def test_restart_after_0001_catches_up_unprocessed_day(self):
        now = self.scheduled_for.replace(minute=10)

        result = schedule.next_event_at(now, self.settings)

        self.assertEqual(result, self.scheduled_for)

    def test_event_journal_or_watermark_prevents_restart_duplicate(self):
        tomorrow = self.scheduled_for.replace(day=31)
        now = self.scheduled_for.replace(minute=10)
        event_id, _ = runner.scheduled_event_identity(
            self.scheduled_for
        )
        cases = {
            "journal": (
                runner.event_result_path(self.root, event_id),
                {
                    "schema_version": 1,
                    "event_id": event_id,
                    "scheduled_for": self.scheduled_for.isoformat(),
                    "payload": {
                        "success": False,
                        "error_code": "AUTH_EXPIRED",
                        "run_id": None,
                        "cycle_id": "2026-07-11--2026-08-10",
                        "mode": "current",
                        "started_at": self.scheduled_for.isoformat(),
                        "finished_at": self.scheduled_for.isoformat(),
                        "next_scheduled_for": None,
                    },
                },
            ),
            "watermark": (
                self.watermark,
                {
                    "event_id": event_id,
                    "local_date": "2026-07-30",
                    "outcome": "success",
                },
            ),
        }
        for name, (path, payload) in cases.items():
            with self.subTest(name=name):
                path.parent.mkdir(parents=True, exist_ok=True)
                path.write_text(json.dumps(payload), encoding="utf-8")
                self.assertEqual(
                    schedule.next_event_at(now, self.settings),
                    tomorrow,
                )
                path.unlink()

    def test_transient_failure_waits_for_backoff_then_retries(self):
        with patch.object(
            schedule.runner,
            "run_once",
            side_effect=[
                {"success": False, "error_code": "LOCKED"},
                {"success": True, "error_code": ""},
            ],
        ) as run_once, patch.object(
            schedule,
            "RETRY_BASE_SECONDS",
            "60",
        ), patch.object(schedule, "RETRY_MAX_SECONDS", "900"):
            self._request()
            first = schedule.run_signal_if_due(
                self.settings,
                now=self.scheduled_for,
            )
            early = schedule.run_signal_if_due(
                self.settings,
                now=self.scheduled_for.replace(second=59),
            )
            retry_payload = json.loads(
                self.signal.read_text(encoding="utf-8")
            )
            retried = schedule.run_signal_if_due(
                self.settings,
                now=self.scheduled_for.replace(minute=2),
            )

        self.assertFalse(first)
        self.assertIsNone(early)
        self.assertTrue(retried)
        self.assertEqual(run_once.call_count, 2)
        self.assertEqual(retry_payload["attempt"], 1)
        self.assertEqual(
            retry_payload["next_attempt_at"],
            "2026-07-30T00:02:00-03:00",
        )
        self.assertFalse(self.signal.exists())
        self.assertEqual(
            json.loads(self.watermark.read_text(encoding="utf-8"))[
                "outcome"
            ],
            "success",
        )

    def test_permanent_failure_finishes_event_without_retry(self):
        with patch.object(
            schedule.runner,
            "run_once",
            return_value={
                "success": False,
                "error_code": "AUTH_EXPIRED",
            },
        ) as run_once:
            self._request()
            result = schedule.run_signal_if_due(
                self.settings,
                now=self.scheduled_for,
            )
            repeated = schedule.run_signal_if_due(
                self.settings,
                now=self.scheduled_for.replace(hour=1),
            )

        watermark = json.loads(
            self.watermark.read_text(encoding="utf-8")
        )
        self.assertFalse(result)
        self.assertIsNone(repeated)
        self.assertEqual(run_once.call_count, 1)
        self.assertFalse(self.signal.exists())
        self.assertEqual(watermark["outcome"], "terminal")
        self.assertEqual(watermark["error_code"], "AUTH_EXPIRED")

    def test_active_claim_is_not_recovered_or_dispatched(self):
        claim = self._claim("a" * 32)
        event_id, _ = runner.scheduled_event_identity(
            self.scheduled_for
        )
        owner_lock = self.runtime / "events" / f"{event_id}.lock"
        with file_lock(owner_lock, wait_seconds=0), patch.object(
            schedule.runner,
            "run_once",
        ) as run_once:
            result = schedule.run_signal_if_due(
                self.settings,
                now=self.scheduled_for.replace(minute=10),
            )

        self.assertIsNone(result)
        self.assertTrue(claim.exists())
        run_once.assert_not_called()

    def test_orphaned_claim_is_recovered_after_restart(self):
        claim = self._claim("b" * 32)
        with patch.object(
            schedule.runner,
            "run_once",
            return_value={"success": True, "error_code": ""},
        ) as run_once:
            result = schedule.run_signal_if_due(
                self.settings,
                now=self.scheduled_for.replace(minute=10),
            )

        self.assertTrue(result)
        self.assertFalse(claim.exists())
        run_once.assert_called_once()
        self.assertIs(
            run_once.call_args.kwargs["settings"],
            self.settings,
        )

    def test_invalid_runner_config_is_terminal_and_sanitized(self):
        secret = "https://user:password@example.invalid/?token=secret"
        self._request()
        with patch.object(
            schedule.runner,
            "run_once",
            side_effect=ValueError(secret),
        ), patch.object(schedule.logger, "warning") as warning:
            result = schedule.run_signal_if_due(
                self.settings,
                now=self.scheduled_for,
            )

        watermark = json.loads(
            self.watermark.read_text(encoding="utf-8")
        )
        logged = " ".join(str(call) for call in warning.call_args_list)
        self.assertFalse(result)
        self.assertFalse(self.signal.exists())
        self.assertEqual(watermark["error_code"], "CONFIG_INVALID")
        self.assertIn("CONFIG_INVALID", logged)
        self.assertNotIn(secret, logged)


class FinanceiroMedicaoWorkerIntegrationTests(unittest.TestCase):
    def test_overdue_event_runs_after_longer_scheduled_job(self):
        current = [datetime(2026, 7, 30, 23, 50)]
        multiplica_at = current[0]
        financeiro_at = datetime(2026, 7, 31, 0, 1)

        def finish_multiplica_after_deadline():
            current[0] = datetime(2026, 7, 31, 0, 2)
            return True

        with patch.object(worker, "_ensure_dirs"), patch.object(
            worker,
            "_scheduled_events",
            return_value=[
                ("multiplica", multiplica_at),
                ("financeiro_medicao", financeiro_at),
            ],
        ) as scheduled_events, patch.object(
            worker,
            "_local_now",
            side_effect=lambda: current[0],
        ), patch.object(
            worker,
            "_run_scheduled_multiplica",
            side_effect=finish_multiplica_after_deadline,
        ), patch.object(
            worker,
            "_run_scheduled_financeiro_medicao",
            side_effect=KeyboardInterrupt(),
        ) as financeiro:
            worker.loop_forever()

        financeiro.assert_called_once_with(scheduled_for=financeiro_at)
        scheduled_events.assert_called_once_with()

    def test_disabled_schedule_ignores_invalid_config_on_import(self):
        environment = os.environ.copy()
        environment.update(
            {
                "FINANCEIRO_MEDICAO_SCHEDULE_ENABLED": "false",
                "FINANCEIRO_MEDICAO_TIMEZONE": "Invalid/Timezone",
                "FINANCEIRO_MEDICAO_SCHEDULE_HOUR": "not-an-hour",
                "FINANCEIRO_MEDICAO_SCHEDULE_MINUTE": "not-a-minute",
            }
        )
        completed = subprocess.run(
            [sys.executable, "-c", "import worker; print('worker-imported')"],
            cwd=Path(worker.__file__).parent,
            env=environment,
            capture_output=True,
            text=True,
            timeout=30,
        )

        self.assertEqual(completed.returncode, 0, msg=completed.stderr)
        self.assertIn("worker-imported", completed.stdout)

    def test_invalid_enabled_config_preserves_other_events(self):
        now = datetime(2026, 7, 30, 10, 1)
        with patch.object(
            worker, "FINANCEIRO_MEDICAO_SCHEDULE_ENABLED", True
        ), patch.object(
            worker, "FINANCEIRO_MEDICAO_TIMEZONE", "Invalid/Timezone"
        ), patch.object(
            worker, "FINANCEIRO_MEDICAO_SCHEDULE_HOUR", "not-an-hour"
        ), patch.object(worker, "PROTHEUS_ENABLED", False), patch.object(
            worker, "ROUTERBOX_ENABLED", True
        ), patch.object(
            worker, "MULTIPLICA_SCHEDULE_ENABLED", True
        ), patch.object(worker.logger, "error") as error:
            names = {name for name, _ in worker._scheduled_events(now)}

        self.assertEqual(names, {"routerbox", "multiplica"})
        logged = " ".join(str(call) for call in error.call_args_list)
        self.assertIn("CONFIG_INVALID", logged)
        self.assertNotIn("Invalid/Timezone", logged)

    def test_financeiro_timezone_does_not_change_existing_schedules(self):
        utc = ZoneInfo("UTC")
        protheus_now = datetime(2026, 7, 30, 5, tzinfo=utc)
        other_now = datetime(2026, 7, 30, 10, 1, tzinfo=utc)
        with patch.object(
            worker, "FINANCEIRO_MEDICAO_TIMEZONE", "Asia/Tokyo"
        ), patch.object(worker, "SCHEDULE_HOUR", 6), patch.object(
            worker, "SCHEDULE_MINUTE", 0
        ), patch.object(
            worker, "MULTIPLICA_SCHEDULE_HOUR", 23
        ), patch.object(
            worker, "MULTIPLICA_SCHEDULE_MINUTE", 50
        ), patch.object(worker, "ROUTERBOX_INTERVAL_MIN", 30):
            protheus = worker._next_run_at(protheus_now)
            multiplica = worker._next_multiplica_run_at(other_now)
            routerbox = worker._next_routerbox_run_at(other_now)

        self.assertEqual(
            protheus, datetime(2026, 7, 30, 6, tzinfo=utc)
        )
        self.assertEqual(
            multiplica, datetime(2026, 7, 30, 23, 50, tzinfo=utc)
        )
        self.assertEqual(
            routerbox, datetime(2026, 7, 30, 10, 30, tzinfo=utc)
        )

    def test_worker_uses_financeiro_catch_up_when_enabled(self):
        timezone = ZoneInfo("America/Sao_Paulo")
        now = datetime(2026, 7, 30, 0, 10, tzinfo=timezone)
        with tempfile.TemporaryDirectory() as temporary, patch.dict(
            os.environ,
            {
                "FINANCEIRO_MEDICAO_LOGA_URL": (
                    "https://dashboard.loga.net.br/medicao_pagamento"
                ),
                "FINANCEIRO_MEDICAO_RUNTIME_ROOT": str(
                    Path(temporary) / "financeiro_medicao"
                ),
                "FINANCEIRO_MEDICAO_SCHEDULE_ENABLED": "true",
                "FINANCEIRO_MEDICAO_SCHEDULE_HOUR": "0",
                "FINANCEIRO_MEDICAO_SCHEDULE_MINUTE": "1",
                "FINANCEIRO_MEDICAO_TIMEZONE": "America/Sao_Paulo",
            },
            clear=True,
        ), patch.object(
            worker, "FINANCEIRO_MEDICAO_SCHEDULE_ENABLED", True
        ), patch.object(
            worker, "FINANCEIRO_MEDICAO_TIMEZONE", "America/Sao_Paulo"
        ), patch.object(
            worker, "FINANCEIRO_MEDICAO_SCHEDULE_HOUR", 0
        ), patch.object(
            worker, "FINANCEIRO_MEDICAO_SCHEDULE_MINUTE", 1
        ), patch.object(
            worker,
            "FINANCEIRO_MEDICAO_RUNTIME_ROOT",
            Path(temporary) / "financeiro_medicao",
            create=True,
        ), patch.object(worker, "PROTHEUS_ENABLED", False), patch.object(
            worker, "ROUTERBOX_ENABLED", False
        ), patch.object(worker, "MULTIPLICA_SCHEDULE_ENABLED", False):
            events = dict(worker._scheduled_events(now))

        self.assertEqual(
            events["financeiro_medicao"],
            datetime(2026, 7, 30, 0, 1, tzinfo=timezone),
        )
