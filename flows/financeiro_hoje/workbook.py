"""Validacao e consolidacao fiel dos relatorios XLSX do financeiro."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
import hashlib
from pathlib import Path
import zipfile
from xml.etree.ElementTree import ParseError

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException


SHEET_NAME = "Consulta"
HEADER_ALIASES = {
    "DataAbertura": "Data_Abertura",
    "HoraAbertura": "Hora_Abertura",
    "Agendamento": "Data_Hora",
    "DesignadoPara": "Designado_Para",
    "Causa": "Causa OS",
    "GrupoCobranca": "Grupo Cobran\u00e7a",
    "CausaAtend": "Causa Atend",
}
REQUIRED_HEADERS = ("Numero", "SituacaoOS", "Fluxo")


class WorkbookInvalid(ValueError):
    """Indica que um relatorio nao atende ao contrato de entrada."""


@dataclass(frozen=True)
class WorkbookArtifact:
    path: Path
    headers: tuple[str, ...]
    rows: int
    size: int
    sha256: str


def validate(path: str | Path, created_after: datetime | None = None) -> WorkbookArtifact:
    """Valida integralmente um XLSX e devolve seus metadados apos fecha-lo."""
    file_path = Path(path)
    if not file_path.is_file():
        raise WorkbookInvalid(f"arquivo inexistente: {file_path}")
    if created_after is not None and file_path.stat().st_mtime < created_after.timestamp():
        raise WorkbookInvalid("arquivo anterior ao inicio da execucao")
    if not zipfile.is_zipfile(file_path):
        raise WorkbookInvalid("arquivo nao e um ZIP/OpenXML valido")

    try:
        with file_path.open("rb") as source:
            try:
                book = load_workbook(source, read_only=True, data_only=False)
            except (InvalidFileException, KeyError, OSError, ParseError, ValueError, zipfile.BadZipFile) as exc:
                exc.__traceback__ = None
                raise WorkbookInvalid("arquivo nao e um ZIP/OpenXML valido") from None

            try:
                if SHEET_NAME not in book.sheetnames:
                    raise WorkbookInvalid(f"aba obrigatoria ausente: {SHEET_NAME}")

                sheet = book[SHEET_NAME]
                rows = sheet.iter_rows(values_only=True)
                try:
                    headers = tuple(next(rows))
                except StopIteration as exc:
                    raise WorkbookInvalid("planilha sem cabecalhos") from exc

                missing = [header for header in REQUIRED_HEADERS if header not in headers]
                if missing:
                    raise WorkbookInvalid(f"cabecalhos obrigatorios ausentes: {', '.join(missing)}")

                row_count = sum(1 for row in rows if any(cell is not None for cell in row))
                if row_count == 0:
                    raise WorkbookInvalid("planilha sem linhas")
            finally:
                book.close()
    except WorkbookInvalid:
        raise
    except (OSError, ParseError, ValueError, zipfile.BadZipFile) as exc:
        raise WorkbookInvalid("falha ao ler workbook por completo") from exc

    return WorkbookArtifact(
        path=file_path,
        headers=headers,
        rows=row_count,
        size=file_path.stat().st_size,
        sha256=_sha256(file_path),
    )


def consolidate(loga: str | Path, acerta: str | Path, output: str | Path) -> WorkbookArtifact:
    """Concatena LOGA e ACERTA, sem transformar ou deduplicar qualquer linha."""
    loga_artifact = validate(loga)
    acerta_artifact = validate(acerta)
    loga_headers = _canonical_headers(loga_artifact.headers)
    acerta_headers = _canonical_headers(acerta_artifact.headers)
    if loga_headers != acerta_headers:
        raise WorkbookInvalid("cabecalhos das fontes sao diferentes")

    output_path = Path(output)
    book = Workbook(write_only=True)
    sheet = book.create_sheet(SHEET_NAME)
    sheet.append(loga_headers)
    try:
        _copy_rows(loga_artifact.path, sheet)
        _copy_rows(acerta_artifact.path, sheet)
        book.save(output_path)
    finally:
        book.close()

    return validate(output_path)


def _canonical_headers(headers: tuple[str, ...]) -> tuple[str, ...]:
    return tuple(HEADER_ALIASES.get(header, header) for header in headers)


def _copy_rows(path: Path, target) -> None:
    source = load_workbook(path, read_only=True, data_only=False)
    try:
        rows = source[SHEET_NAME].iter_rows(values_only=True)
        next(rows)
        for row in rows:
            target.append(row)
    finally:
        source.close()


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
