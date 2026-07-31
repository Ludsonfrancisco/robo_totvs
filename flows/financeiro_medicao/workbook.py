from dataclasses import dataclass
from datetime import date, datetime
from contextlib import nullcontext
from pathlib import Path
from xml.etree.ElementTree import ParseError
from zipfile import BadZipFile, ZipFile
from zlib import error as ZlibError

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


SHEET_NAME = "Base Medição de Pagamento"
REQUIRED_HEADERS = (
    "Numero",
    "DataAbertura",
    "Protocolo",
    "Cliente",
    "Nome",
    "Cidade",
    "Bairro",
    "FimUsuario",
    "Executor",
    "FimData",
    "RazaoSocial",
    "NomeFantasia",
    "Fluxo",
    "Fluxo1",
    "Topico",
    "Causa",
    "TipoEncerramento",
    "Expurgado",
    "Motivo_Expurgo",
    "Valor Atividade",
)
MAX_WORKBOOK_BYTES = 256 * 1024 * 1024
_FIM_DATA_INDEX = REQUIRED_HEADERS.index("FimData")


class WorkbookInvalid(RuntimeError):
    pass


@dataclass(frozen=True)
class WorkbookInfo:
    row_count: int
    headers: tuple[str, ...]
    size: int


def _as_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for pattern in ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S"):
            try:
                return datetime.strptime(value, pattern).date()
            except ValueError:
                continue
    return None


def _validate_archive(stream):
    """Reject duplicate ZIP names, which make an XLSX package ambiguous."""
    with ZipFile(stream) as archive:
        names = set()
        for entry in archive.infolist():
            if entry.filename in names:
                raise WorkbookInvalid("Arquivo de medição inválido.")
            names.add(entry.filename)


def validate_workbook(path_or_stream, query_start, query_end):
    start = _as_date(query_start)
    end = _as_date(query_end)
    if start is None or end is None or start > end:
        raise WorkbookInvalid("Período de consulta inválido.")

    is_stream = all(
        callable(getattr(path_or_stream, name, None))
        for name in ("read", "seek", "tell")
    )
    if is_stream:
        stream = path_or_stream
        try:
            stream.seek(0, 2)
            size = stream.tell()
            stream.seek(0)
        except (OSError, ValueError, TypeError) as error:
            raise WorkbookInvalid(
                "Arquivo de medição inválido."
            ) from error
        stream_context = nullcontext(stream)
    else:
        workbook_path = Path(path_or_stream)
        try:
            size = workbook_path.stat().st_size
        except OSError as error:
            raise WorkbookInvalid(
                "Arquivo de medição inválido."
            ) from error
        if not workbook_path.is_file():
            raise WorkbookInvalid("Arquivo de medição inválido.")
        stream_context = workbook_path.open("rb")

    if not 0 < size <= MAX_WORKBOOK_BYTES:
        raise WorkbookInvalid("Arquivo de medição inválido.")

    workbook = None
    rows = None
    try:
        with stream_context as stream:
            _validate_archive(stream)
            stream.seek(0)
            workbook = load_workbook(stream, read_only=True, data_only=True)
            if SHEET_NAME not in workbook.sheetnames:
                raise WorkbookInvalid("Planilha de medição inválida.")

            worksheet = workbook[SHEET_NAME]
            rows = worksheet.iter_rows(values_only=True)
            try:
                headers = next(rows)
            except StopIteration as error:
                raise WorkbookInvalid("Planilha de medição inválida.") from error

            if headers is None or tuple(headers) != REQUIRED_HEADERS:
                raise WorkbookInvalid("Cabeçalhos da planilha inválidos.")

            row_count = 0
            for row_number, row in enumerate(rows, start=2):
                if not row or all(value is None for value in row):
                    continue
                if len(row) < len(REQUIRED_HEADERS):
                    raise WorkbookInvalid("Linha da planilha inválida.")

                fim_data = _as_date(row[_FIM_DATA_INDEX])
                if fim_data is None:
                    raise WorkbookInvalid("Data FimData inválida.")
                if not start <= fim_data <= end:
                    raise WorkbookInvalid("Data FimData fora do período de consulta.")
                row_count += 1

            if row_count == 0:
                raise WorkbookInvalid("Planilha de medição sem dados.")

            return WorkbookInfo(row_count=row_count, headers=REQUIRED_HEADERS, size=size)
    except WorkbookInvalid:
        raise
    except (
        BadZipFile,
        InvalidFileException,
        OSError,
        ValueError,
        KeyError,
        ParseError,
        RuntimeError,
        TypeError,
        ZlibError,
    ) as error:
        raise WorkbookInvalid("Planilha de medição inválida.") from error
    finally:
        if rows is not None:
            close = getattr(rows, "close", None)
            if close is not None:
                close()
        if workbook is not None:
            workbook.close()
